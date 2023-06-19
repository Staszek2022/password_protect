import os
import sys
import tkinter
import hashlib
import openpyxl
import subprocess
from tkinter import messagebox
import sv_ttk
import glob

PATH_TO_FILE = "Zeszyt1.xlsx" # UWAGA: w miejscu jednego "\" trzeba dać dwa "\\"
LETTER_SERVER = "m:" # UWAGA: musi być dwukropek

isMounted = False
# PATH_TO_TEMP = sys._MEIPASS + "\\" # UWAGA: na kńcu musi być "\\" 
PATH_TO_TEMP = "temppy\\"


# otwórz plik z hasłami i pozyskaj arkusze
zeszyt = openpyxl.load_workbook(PATH_TO_FILE)
arkusze = zeszyt.sheetnames


# funkcja do pozyskiania słownika z arkusza Excela(to czyta od drugiego wiersza)
def to_dict(arkusz):
    return {str(row[0].value): str(row[1].value) for row in arkusz.iter_rows(min_row=2, min_col=1, max_col=2)}

# funkcja do mapowania dysku zdalnego(UWAGA: usuwa poświadczenia, bez loginu i hasła w drugiej próbie nie będzie działć)
def mapNetworkDrives(path, letter):
    global isMounted
    # usunięcie dysku z podanej litery
    subprocess.call(r'net use '+ letter + r' /delete /y', shell=True)

    # jeżeli w pliku jest login i hasło do serwera
    if ustawienia["server username"] == "None":
        isMounted = True
        subprocess.call(r'net use m: ' + path + r' /user:'+ ustawienia["server username"] + ustawienia["server password"], shell=True)

    else:
        # jeśli nie
        print(r'net use m: ' + path)


# funkcja uruchamiająca się przy zamknięciu

def on_closing():
    if isMounted:
        if messagebox.askokcancel("Wyjście", "Czy na pewno chcesz wyjść?\nUWAGA: po zamknięciu nastąpi wylogowanie."):
            subprocess.call( r'net use '+ LETTER_SERVER + r' /delete /y')
    login_screen.destroy()


# funkcja do dodawania zmmienionych haseł
def update_list(slownik_klasy, klasa):
    dir_path = f"{os.path.join(ustawienia['folders path'], klasa)}\\*\\passwords.txt"
    for file in glob.glob(dir_path, recursive=True):
        with open(file) as my_file:
            if slownik_klasy[file.split("\\")[-2]] == my_file.readline():
                slownik_klasy[file.split("\\")[-2]] = my_file.readline()


# pozyskaj ustawienia
ustawienia = to_dict(zeszyt['ustawienia'])

# definiowanie funkcji logoania ucznia
def login2(klasa, uczen, haslo):
    
    # tworzenie słownika z hasłami danej klasy pobranymi z Excela
    uczniowie_klasy = to_dict(zeszyt[klasa])
    # aktualizowanie słownika
    uczniowie_klasy = update_list(uczniowie_klasy)
    
    # szyfrowanie hasła przzez sha256()
    hashed_password = hashlib.sha256(haslo.encode('utf-8')).hexdigest()

    # Czy ten uczeń jest zapisany w pliku?
    if uczen in uczniowie_klasy:
        # Czy hasło jest pranidłowe?
        if haslo == uczniowie_klasy[uczen] or hashed_password == uczniowie_klasy[uczen]:
            # jeśli tak zaloguj
            if not isChangePassword:
                path_to_student_file = os.path.join(ustawienia["folders path"], klasa, uczen)
                print(path_to_student_file)
                mapNetworkDrives(path_to_student_file, LETTER_SERVER)

            # ta funkcja po przerobieniu mogłaby zmieniać hasła
            else:
                changePassword(klasa, uczniowie_klasy, uczen, haslo, hashed_password)
        else:
            message.set("Nieprawidłowe hasło!!!")
            messagebox.showerror("Error", "Nieprawidłowe hasło!!!")

    # Jeśli nie istnieje napisz to.
    else:
        message.set("Taki login nie istnieje!!!")
        messagebox.showerror("Error", "Taki login nie istnieje!!!")


# definioanie ogólnej funkcji logowania
def login(isChangePassword1=False):
    # pozyskianie danych do zmiennych
    uname = username.get()
    print(uname)
    pwd = password.get()
    print(pwd)
    klasa = clicked.get()
    print(klasa)
    global isChangePassword
    isChangePassword = isChangePassword1

    # Czy loguje się admin?
    if uname == ustawienia["admin login"] and pwd == ustawienia["admin password"]:
        AdminLogin()
    # Czy pole jest puste?
    elif uname == "" and pwd == "":
        message.set("Nic nie wwpisałeś/aś!!!")
    else:
        # jeśli nie zaloguj się jako uczeń
        login2(klasa, uname, pwd)




# funkcja do zmiany hasła po pierwszym kliknięciu
def changePassword(klasa, uczniowie_klasy, uczen, haslo, hashed_password):
    # Czy ktoś już zmieniał hasło
    if not os.path.isfile(os.path.join(ustawienia["folders path"], klasa, uczen, "passwords.txt")):

        # ususwanie tekstu z pola do wpisywania hasła
        username.set("")
        password.set("")

        # ustawianie noej treści podpowiedzi
        login_help.set("nowe hasło")
        password_help.set("jeszcze raz")

        # usuwanie przycisku i menu klasy
        password_change.destroy()
        menu.destroy()

        login_button.config(text="Zatwierdź", # zmiana przycisku login na zatwierdź
                            command=lambda: changePassword2(klasa, uczniowie_klasy, uczen, haslo, hashed_password))  # przekierowanie do funkcji changepassword2 po kliknięciu

    else:
        messagebox.showerror("Error", "Już zmieniłeś hasło.\nSkontaktuj się z administratorem i poproś o aktualizacje bazy haseł")

def changePassword2(klasa, uczniowie_klasy, uczen, haslo, hashed_password):
    # # Czy ktoś nie wpisał źle hasła
    if username == password: # username to teraz pole 1 do hasla
        hashed_new_password = hashlib.sha256(username.encode('utf-8')).hexdigest()
        f = open(PATH_TO_TEMP + uczen + ".txt", "w")
        f.write(f"\n{hashed_password}\n{hashed_new_password}")
        f.close()
        os.system(f"copy {PATH_TO_TEMP}{uczen}.txt {os.path.join(ustawienia['folders path'], klasa, uczen, 'passwords.txt')}")
        messagebox.showinfo("Ok", "Pomyślnie zmieniono hasło")



def AdminLogin():
    pass


# defining loginform function


def Loginform():
    global login_screen
    login_screen = tkinter.Tk()
    # tytuł strony
    login_screen.title("Logowanie")
    # rozmiar strony
    login_screen.minsize(300, 250)
    login_screen.maxsize(300, 250)
    # dodawanie zmiennych
    global message
    global username
    global password
    global clicked
    global arkusze
    global password_help
    global login_help
    global password_change
    global login_button
    global menu
    username = tkinter.StringVar()
    password = tkinter.StringVar()
    message = tkinter.StringVar()
    clicked = tkinter.StringVar(login_screen)
    clicked.set(arkusze[1])
    password_help = tkinter.StringVar()
    password_help.set("Hasło * ")
    login_help = tkinter.StringVar()
    login_help.set("Login * ")
    # Creating layout of login form
    tkinter.Label(login_screen, width="300", text="Podaj swoje dane", bg="orange", fg="white").pack()
    # Username Label
    tkinter.Label(login_screen, textvariable=login_help).place(x=20, y=40)
    # Username textbox
    tkinter.Entry(login_screen, textvariable=username).place(x=90, y=42)
    # Password Label
    tkinter.Label(login_screen, textvariable=password_help).place(x=20, y=80)
    # Password textbox
    tkinter.Entry(login_screen, textvariable=password, show="*").place(x=90, y=82)
    # textbox klasy
    tkinter.Label(login_screen, text="Klasa * ").place(x=20, y=128)
    # menu klasy
    menu = tkinter.OptionMenu(login_screen, clicked, *arkusze[1:])
    menu.config(indicatoron=True)
    menu.place(x=90, y=122)
    # Label for displaying login status[success/failed]
    tkinter.Label(login_screen, text="", textvariable=message).place(x=95, y=160)
    # Login button
    login_button = tkinter.Button(login_screen, text="Login", width=10, height=1, bg="orange", command=login)
    login_button.place(x=105, y=190)
    # Przycisk do zmiany hasłą
    password_change = tkinter.Button(login_screen, text="Zmień hasło", font=("Segoe UI",8), width=10, height=1, command=lambda: login(True))
    password_change.place(x=108, y=220)
    

    login_screen.protocol("WM_DELETE_WINDOW", on_closing)
    sv_ttk.set_theme("light")
    login_screen.mainloop()


# calling function Loginform
Loginform()
